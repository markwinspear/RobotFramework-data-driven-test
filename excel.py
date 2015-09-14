"""Functions which utilise the excel library"""
from openpyxl import load_workbook

logins = {}
data = []

def get_logins(filename, sheetname):
    """Returns usernames and passwords when given the filename and sheetname for the test data"""
    wk = load_workbook(filename)
    ws = wk.get_sheet_by_name(sheetname)
    rowcount = 2            # assumes 1st row contains column headers
    for r in ws.rows:
        username = ws.cell(row = rowcount, column = 2)
        password = ws.cell(row = rowcount, column = 3)
       # expected_result = ws.cell(row = rowcount, column = 4)
        print(username.value)
        print (password.value)
        #print (expected_result.value)
        rowcount = rowcount + 1
        logins[username.value] = [password.value]
     #   logins[username.value] = [password.value, expected_result.value]
        try:                            #picks up on an empty row on the end - this removes it
            del logins[None]
        except KeyError:
            pass
    return logins

#TODO: Make expected result load in from sheet using http://stackoverflow.com/questions/20585920/how-to-add-multiple-values-to-a-dictionary-key-in-python...
